import plotly.graph_objs as go
import pandas as pd
import matplotlib
from parest_copasi import parameter_estimation_online
from plotly import tools
from models import batch_model_mu
from openpyxl import load_workbook
matplotlib.use('TkAgg')


def monitoring(online_data,filename_experimental_data1,filename_experimental_data2,alpha_lower_bound,
               alpha_upper_bound, beta_lower_bound,beta_upper_bound):


    # Calculate the difference in time, so we can select all the data that corresponds to 1 reactor

    time = pd.to_timedelta(online_data['Time      '])
    shifted_time = time.shift(periods=-1)
    delta = shifted_time - time
    online_data['delta'] = delta

    # Select the rows with difference in time between 46 and 47 minutes
    # and create new dataframe that we will be working with
    selected_data = online_data[(online_data['delta'] >= '00:46:00') & (online_data['delta'] <= '00:47:00')]

    # Calculation of the CO2 evolution rate
    CER = selected_data['CO2 (Vol.%)'] * 10 - 0.04 * 10  # unit [(mol_co2/mol_totalgas)/min] / [%CO2/min]

    # Reset the selected time so it starts from time = 0, convert it and then use it to calculate tCER
    selected_time = pd.to_timedelta(selected_data['Time      '])
    selected_time.reset_index(inplace=True, drop=True)
    reset_selected_time = selected_time - selected_time[0]
    selected_datetimes = pd.to_datetime(reset_selected_time)
    selected_time = selected_datetimes.dt.time

    # convert time to decimals and in minutes
    selected_time_decimals = pd.DataFrame(columns=['Time'])
    for i in range(0, len(selected_time)):
        h = selected_time[i].strftime('%H')
        m = selected_time[i].strftime('%M')
        s = selected_time[i].strftime('%S')
        result = int(h) * 60 + int(m) + int(s) / 60.0  # [min]
        selected_time_decimals.loc[
            i, ['Time']] = result  # This puts the results in the iterated indexes in the Time column

    # Calculate tCER
    CER.reset_index(inplace=True, drop=True)

    selected_time_decimals = selected_time_decimals.iloc[:, 0]

    tCER = []
    tCER.append(0)  # Here set the initial value of tCER if we have that.

    for i in range(0, (len(selected_time_decimals) - 1)):

        tCER_i = ((CER[i] + CER[i + 1]) / 2) * (selected_time_decimals[i + 1] - selected_time_decimals[i]) + tCER[i]
        tCER.append(tCER_i)

    mu = CER / tCER  # [1/min]
    print(mu)
    mu = (mu*60)  # [1/h]
    print(mu)

    selected_time_decimals_hours = selected_time_decimals / 60

    r = batch_model_mu()
    r.timeCourseSelections = ['time', 'glucose', 'serine', 'biomass', 'mu']

    start_time = selected_time_decimals_hours[0]
    end_time = selected_time_decimals_hours.iloc[1]
    results = r.simulate(start_time, end_time, 2)

    # we probably did'nt have to simulate since we just want the first row.
    # But it makes it easier to make the data frame
    initial_values = results[0:1]
    data_frame = pd.DataFrame(initial_values)
    data_frame.columns = ['time', 'glucose', 'serine', 'biomass', 'mu']

    # It could be an idea to have this as a temporary file.
    wb = load_workbook("/Users/s144510/Documents/fermentationtool/output.xlsx")
    ws = wb['Sheet1']
    row = list(data_frame.iloc[-1])
    ws.append(row)
    wb.save("/Users/s144510/Documents/fermentationtool/output.xlsx")

    output_values = '/Users/s144510/Documents/fermentationtool/output.xlsx'
    output_values = pd.ExcelFile(output_values)

    # Loads the sheet we want to work with
    data_frame = output_values.parse('Sheet1')
    data_frame = data_frame.drop_duplicates()

    writer = pd.ExcelWriter('/Users/s144510/Documents/fermentationtool/mu.xlsx')
    mu.to_excel(writer, 'Sheet1', index=False)
    writer.save()

    if (len(mu) - len(data_frame['mu'])) > 1:

        for i in range(0, (len(mu) - 1)):
            r.reset()
            r.mu = mu[i + 1]
            glucose = data_frame['glucose'].iloc[-1]
            serine = data_frame['serine'].iloc[-1]
            biomass = data_frame['biomass'].iloc[-1]

            alpha_online, beta_online = parameter_estimation_online(filename_experimental_data1,
                                                                    filename_experimental_data2,
                                                                    alpha_lower_bound, alpha_upper_bound,
                                                                    beta_lower_bound, beta_upper_bound,
                                                                    str(mu[i + 1]), str(glucose), str(serine),
                                                                    str(biomass))
            print(alpha_online,beta_online)
            print(biomass)
            print(glucose)
            print(serine)
            r.glucose = glucose
            r.biomass = biomass
            r.serine = serine
            r.alpha = float(alpha_online)
            r.beta = float(beta_online)
            start_time = selected_time_decimals_hours[i]
            end_time = selected_time_decimals_hours[i + 1]
            results = r.simulate(start_time, end_time, 2)
            simulated_row = results[-1:]
            print(simulated_row)
            new_dataframe = pd.DataFrame(simulated_row)

            wb = load_workbook("/Users/s144510/Documents/fermentationtool/output.xlsx")
            ws = wb['Sheet1']
            row = list(new_dataframe.iloc[-1])
            ws.append(row)
            wb.save("output.xlsx")

            new_dataframe.columns = ['time', 'glucose', 'serine', 'biomass', 'mu']
            data_frame = data_frame.append(new_dataframe, ignore_index=True)

    elif (len(mu) - len(data_frame['mu'])) == 1:
        r.reset()
        r.mu = mu.iloc[-1]
        glucose = data_frame['glucose'].iloc[-1]
        serine = data_frame['serine'].iloc[-1]
        biomass = data_frame['biomass'].iloc[-1]

        alpha_online, beta_online = parameter_estimation_online(filename_experimental_data1,
                                                                filename_experimental_data2,
                                                                alpha_lower_bound, alpha_upper_bound,
                                                                beta_lower_bound, beta_upper_bound,
                                                                str(mu.iloc[-1]), str(glucose), str(serine),
                                                                str(biomass))
        r.glucose = glucose
        r.biomass = biomass
        r.serine = serine
        r.alpha = float(alpha_online)
        r.beta = float(beta_online)
        start_time = selected_time_decimals_hours.iloc[-2]
        end_time = selected_time_decimals_hours.iloc[-1]
        results = r.simulate(start_time, end_time, 2)
        simulated_row = results[-1:]
        new_dataframe = pd.DataFrame(simulated_row)

        wb = load_workbook("/Users/s144510/Documents/fermentationtool/output.xlsx")
        ws = wb['Sheet1']
        row = list(new_dataframe.iloc[-1])
        ws.append(row)
        wb.save("/Users/s144510/Documents/fermentationtool/output.xlsx")

        new_dataframe.columns = ['time', 'glucose', 'serine', 'biomass', 'mu']
        data_frame = data_frame.append(new_dataframe, ignore_index=True)

    else:
        print('No new data to simulate model with')

    data_frame = data_frame.drop_duplicates()

    writer = pd.ExcelWriter('/Users/s144510/Documents/fermentationtool/data_online_integration.xlsx')
    data_frame.to_excel(writer, 'Sheet1', index=False)
    writer.save()

    trace1 = go.Scatter(
        x=selected_time_decimals_hours,
        y=selected_data['CO2 (Vol.%)'],
        name='CO2',
        mode='markers'
    )

    trace2 = go.Scatter(
        x=selected_time_decimals_hours,
        y=mu,
        name='mu',
        mode='markers'
    )

    trace3 = go.Scatter(
        x=data_frame['time'],
        y=data_frame['mu'],
        name='mu'
    )

    trace4 = go.Scatter(
        x=data_frame['time'],
        y=data_frame['biomass'],
        name='Biomass'
    )

    trace5 = go.Scatter(
        x=data_frame['time'],
        y=data_frame['serine'],
        name='Serine'
    )

    trace6 = go.Scatter(
        x=data_frame['time'],
        y=data_frame['glucose'],
        name='Glucose'
    )

    fig = tools.make_subplots(rows=2, cols=3, subplot_titles=('CO2 online data', 'mu from CO2',
                                                              'mu from model', 'Biomass from model',
                                                              'Serine from model', 'Glucose from model'))

    fig.append_trace(trace1, 1, 1)
    fig.append_trace(trace2, 1, 2)
    fig.append_trace(trace3, 1, 3)
    fig.append_trace(trace4, 2, 1)
    fig.append_trace(trace5, 2, 2)
    fig.append_trace(trace6, 2, 3)

    fig['layout'].update(height=640, width=1260,
                         margin=dict(
                             l=120,
                             r=100,
                             b=100,
                             t=70,
                             pad=2
                         ))

    fig['layout']['yaxis1'].update(showgrid=True, title='CO2 (%)', exponentformat='power', nticks=10,
                                   tickfont=dict(size=10), domain=[0.65, 1])
    fig['layout']['yaxis2'].update(showgrid=True, title='Mu (1/h)', exponentformat='power', nticks=10,
                                   tickfont=dict(size=10), domain=[0.65, 1])
    fig['layout']['yaxis3'].update(showgrid=True, title='Mu (1/h)', exponentformat='power', nticks=10,
                                   tickfont=dict(size=10), domain=[0.65, 1])
    fig['layout']['yaxis4'].update(showgrid=True, title='Biomass (moles)', exponentformat='power', nticks=10,
                                   tickfont=dict(size=10), domain=[0, 0.35])
    fig['layout']['yaxis5'].update(showgrid=True, title='Serine (moles)', exponentformat='power', nticks=10,
                                   tickfont=dict(size=10), domain=[0, 0.35])
    fig['layout']['yaxis6'].update(showgrid=True, title='Glucose (moles)', exponentformat='power', nticks=10,
                                   tickfont=dict(size=10), domain=[0, 0.35])

    fig['layout']['xaxis1'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
                                   domain=[0, 0.27])
    fig['layout']['xaxis2'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
                                   domain=[0.36, 0.63])
    fig['layout']['xaxis3'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
                                   domain=[0.72, 0.99])
    fig['layout']['xaxis4'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
                                   domain=[0, 0.27])
    fig['layout']['xaxis5'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
                                   domain=[0.36, 0.63])
    fig['layout']['xaxis6'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
                                   domain=[0.72, 0.99])

    return fig
