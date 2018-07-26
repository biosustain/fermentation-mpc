import warnings
warnings.filterwarnings("ignore", message="numpy.dtype size changed")
warnings.filterwarnings("ignore", message="numpy.ufunc size changed")
# Supresses the benign warnings from numpy

import datetime
import dash
import dash_core_components as dcc
import dash_html_components as html
import plotly
from dash.dependencies import Input, Output

# pip install pyorbital
from pyorbital.orbital import Orbital

import pandas as pd
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from models import batch_model
from models import fed_batch_model
from experiments import data
from parest_copasi import parameter_estimation
from parest_copasi import parameter_estimation_online
import os
import sys
import time
import datetime
import numpy as np
from plotly import tools
import plotly
import plotly.graph_objs as go
import tellurium as te
from models import batch_model_mu
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows



# Set filename of the two experimental datasets
filename_experimental_data1 = "data/R1_data_in_moles.csv"
filename_experimental_data2 = "data/R2_data_in_moles.csv"

# Experimental data set 1
experimental_data = pd.read_csv(filename_experimental_data1)
experimental_data = data(experimental_data)

# Experimental data set 2
experimental_data2 = pd.read_csv(filename_experimental_data2)
experimental_data2 = data(experimental_data2)


# #  Parameter estimation
#
#
# Set lower and upper bounds
alpha_lower_bound = "0"
alpha_upper_bound = "100"
beta_lower_bound = "0"
beta_upper_bound = "100"
kc_lower_bound = "0"
kc_upper_bound = "100"
mu_max_lower_bound = "0"
mu_max_upper_bound = "100"


watch_file = 'data/MUX_09-03-2018_18-38-27.XLS'
online_data = pd.ExcelFile(watch_file)
online_data = online_data.parse('Sheet1')


# Calculate the difference in time, so we can select all the data that corresponds to 1 reactor
time = pd.to_timedelta(online_data['Time      '])
shifted_time = time.shift(periods=-1)
delta = shifted_time - time
online_data['delta'] = delta

# Select the rows with difference in time between 46 and 47 minutes
# and create new dataframe that we will be working with
selected_data = online_data[(online_data['delta'] >= '00:46:00') & (online_data['delta'] <= '00:47:00')]

# Calculation of the CO2 evolution rate
CER = selected_data['CO2 (Vol.%)'] * 10 - 0.04 * 10  # unit [(mol_co2/mol_totalgas)/min] / [%CO2/min]

# Reset the selected time so it starts from time = 0, convert it and then use it to calculate tCER
selected_time = pd.to_timedelta(selected_data['Time      '])
selected_time.reset_index(inplace=True, drop=True)
reset_selected_time = selected_time - selected_time[0]
selected_datetimes = pd.to_datetime(reset_selected_time)
selected_time = selected_datetimes.dt.time

# convert time to decimals and in minutes
selected_time_decimals = pd.DataFrame(columns=['Time'])
for i in range(0, len(selected_time)):
    h = selected_time[i].strftime('%H')
    m = selected_time[i].strftime('%M')
    s = selected_time[i].strftime('%S')
    result = int(h) * 60 + int(m) + int(s) / 60.0  # [min]
    selected_time_decimals.loc[
        i, ['Time']] = result  # This puts the results in the iterated indexes in the Time column

# Calculate tCER
CER.reset_index(inplace=True, drop=True)

selected_time_decimals = selected_time_decimals.iloc[:, 0]

tCER = []
tCER.append(0)  # Here set the initial value of tCER if we have that.

for i in range(0, (len(selected_time_decimals)-1)): ##HUSK SUM

    tCER_i = ((CER[i] + CER[i+1])/2)*(selected_time_decimals[i+1]-selected_time_decimals[i]) + tCER[i]
    tCER.append(tCER_i)

mu = CER / tCER
#print(mu)

selected_time_decimals_hours = selected_time_decimals / 60

r = batch_model_mu()
r.timeCourseSelections = ['time', 'glucose', 'serine', 'biomass', 'mu']

start_time = selected_time_decimals_hours[0]
end_time = selected_time_decimals_hours.iloc[1]
results = r.simulate(start_time, end_time, 2)

#print(results)

# we probably didnt have to simulate since we just want the first row.
# But it makes it easier to make the dataframe
initial_values = results[0:1]
data_frame = pd.DataFrame(initial_values)
data_frame.columns = ['time', 'glucose', 'serine', 'biomass', 'mu']
print(data_frame)


#writer = pd.ExcelWriter('output.xlsx')
#data_frame.to_excel(writer,'Sheet1')

#print(data_frame)
#print(mu[1:],'MUUUU')


#writer = pd.ExcelWriter('output.xlsx')
#data_frame.to_excel(writer,'Sheet1', index = False)
#writer.save()

wb = load_workbook("output.xlsx")
ws = wb['Sheet1']
row = list(data_frame.iloc[-1])
ws.append(row)
wb.save("output.xlsx")

#
output_values = 'output.xlsx'
output_values = pd.ExcelFile(output_values)


# Loads the sheet we want to work with
data_frame = output_values.parse('Sheet1')

data_frame = data_frame.drop_duplicates()

print(data_frame)
#
#
# #print(data_frame)
# print(data_frame['mu'])
# print(len(data_frame['mu']))
# print(mu)
# print(len(mu))
print(mu)
#print(type(mu))
#print(mu.iloc[-1],'MUUUUUH')

if (len(mu) - len(data_frame['mu'])) > 1:

    for i in range(0, (len(mu) - 1)):
        r.reset()
        r.mu = mu[i+1]
        glucose = data_frame['glucose'].iloc[-1]
        serine = data_frame['serine'].iloc[-1]
        biomass = data_frame['biomass'].iloc[-1]
        print(glucose,serine,biomass)
        alpha_online, beta_online = parameter_estimation_online(filename_experimental_data1,
                                                                filename_experimental_data2,
                                                                alpha_lower_bound, alpha_upper_bound,
                                                                beta_lower_bound, beta_upper_bound,
                                                                str(mu[i+1]), str(glucose), str(serine),
                                                                str(biomass))

        r.glucose = glucose
        r.biomass = biomass
        r.serine = serine
        r.alpha = float(alpha_online)
        r.beta = float(beta_online)
        start_time = selected_time_decimals_hours[i]
        end_time = selected_time_decimals_hours[i + 1]
        results = r.simulate(start_time, end_time, 2)
        #print(results)
        simulated_row = results[-1:]
        #print(simulated_row)

        new_dataframe = pd.DataFrame(simulated_row)

        wb = load_workbook("output.xlsx")
        ws = wb['Sheet1']
        row = list(new_dataframe.iloc[-1])
        ws.append(row)
        wb.save("output.xlsx")


        new_dataframe.columns = ['time', 'glucose', 'serine', 'biomass', 'mu']
        data_frame = data_frame.append(new_dataframe, ignore_index=True)

elif (len(mu) - len(data_frame['mu'])) == 1:

    #for i in range(0, (len(mu) - 1)):
    r.reset()
    r.mu = mu.iloc[-1]
    glucose = data_frame['glucose'].iloc[-1]
    serine = data_frame['serine'].iloc[-1]
    biomass = data_frame['biomass'].iloc[-1]
    print(glucose,serine,biomass)
    alpha_online, beta_online = parameter_estimation_online(filename_experimental_data1,
                                                            filename_experimental_data2,
                                                            alpha_lower_bound, alpha_upper_bound,
                                                            beta_lower_bound, beta_upper_bound,
                                                            str(mu.iloc[-1]), str(glucose), str(serine),
                                                            str(biomass))

    r.glucose = glucose
    r.biomass = biomass
    r.serine = serine
    r.alpha = float(alpha_online)
    r.beta = float(beta_online)
    start_time = selected_time_decimals_hours.iloc[-2]
    end_time = selected_time_decimals_hours.iloc[-1]
    results = r.simulate(start_time, end_time, 2)
    #print(results)
    simulated_row = results[-1:]
    #print(simulated_row)

    new_dataframe = pd.DataFrame(simulated_row)

    wb = load_workbook("output.xlsx")
    ws = wb['Sheet1']
    row = list(new_dataframe.iloc[-1])
    ws.append(row)
    wb.save("output.xlsx")


    new_dataframe.columns = ['time', 'glucose', 'serine', 'biomass', 'mu']
    data_frame = data_frame.append(new_dataframe, ignore_index=True)

else:
    print('No new data to simulate model with')


data_frame = data_frame.drop_duplicates()
print(data_frame)



# wb = load_workbook("output.xlsx")
# ws = wb['Sheet1']
# row = list(data_frame)
# ws.append(row)
# wb.save("output.xlsx")
# print(data_frame)



# print(data_frame)
# print(mu)

#
#
#
#
#
#
#
#
#
#     print(data_frame, "on the top")
#     #print(results,"on the top")
#     #print(data_frame['glucose'].iloc[-1])
#     #r.reset()
#     print(mu[i+1],'mu')
#     r.mu = mu[i+1]
#     glucose = data_frame['glucose'].iloc[-1]
#     serine = data_frame['serine'].iloc[-1]
#     biomass = data_frame['biomass'].iloc[-1]
#     print(glucose,serine,biomass)
#
#     alpha_online, beta_online = parameter_estimation_online(filename_experimental_data1,
#                                                             filename_experimental_data2,
#                                                             alpha_lower_bound, alpha_upper_bound,
#                                                             beta_lower_bound, beta_upper_bound,
#                                                             str(mu[i+1]), str(glucose), str(serine),
#                                                             str(biomass))
#     #print((mu[i+1]))
#     print(alpha_online,beta_online)
#     r.alpha = float(alpha_online)
#     r.beta = float(beta_online)
#     start_time = selected_time_decimals_hours[i]
#     end_time = selected_time_decimals_hours[i + 1]
#     results = r.simulate(start_time, end_time, 2)
#     print(results)
#     simulated_row = results[-1:]
#     print(simulated_row)
#     new_dataframe = pd.DataFrame(simulated_row)
#     new_dataframe.columns = ['time', 'glucose', 'serine', 'biomass', 'mu']
#     data_frame = data_frame.append(new_dataframe, ignore_index=True)
#     print(data_frame)
#     #print("ROUND",i)
#
# print(data_frame)









#
# # if np.isnan(mu[0]) == True:
# #     print('Needs more time points to simulate data')
# #
# # else:
#
selected_time_decimals_hours = selected_time_decimals / 60
#
#r = batch_model_mu()
#r.mu = mu[0]
#r.timeCourseSelections = ['time', 'glucose', 'serine', 'biomass', 'mu']

#start_time = selected_time_decimals_hours[0]
#end_time = selected_time_decimals_hours.iloc[1]
#results = r.simulate(start_time, end_time, 2)

#print(results)

# # we probably didnt have to simulate since we just want the first row.
# # But it makes it easier to make the dataframe
#initial_values = results[0:1]
#data_frame = pd.DataFrame(initial_values)
#data_frame.columns = ['time', 'glucose', 'serine', 'biomass', 'mu']
#print(data_frame)
# #data_frame['CO2 (%)'] = selected_data['CO2 (Vol.%)'].iloc[0]
#
# print(data_frame)
#
# r.mu = mu.iloc[-2]
# print(mu.iloc[-2])
# glucose = data_frame['glucose'].iloc[-1]
# serine = data_frame['serine'].iloc[-1]
# biomass = data_frame['biomass'].iloc[-1]
#
#
#
#
#
#
#
# # r.reset()
# # print(mu.iloc[-2])
# # r.mu = mu.iloc[-2]
# # glucose = data_frame['glucose'].iloc[-1]
# # serine = data_frame['serine'].iloc[-1]
# # biomass = data_frame['biomass'].iloc[-1]
# print(glucose, serine, biomass, "SPECIES")
# #
# alpha_online, beta_online = parameter_estimation_online(filename_experimental_data1,
#                                                         filename_experimental_data2,
#                                                         alpha_lower_bound, alpha_upper_bound,
#                                                         beta_lower_bound, beta_upper_bound,
#                                                         str(mu.iloc[-2]), str(glucose), str(serine),
#                                                         str(biomass))
# # print(alpha_online,beta_online)
# r.alpha = float(alpha_online)
# r.beta = float(beta_online)
# start_time = selected_time_decimals_hours.iloc[-3]
# end_time = selected_time_decimals_hours.iloc[-2]
# results = r.simulate(start_time, end_time, 2)
# # print(results,'hello')
# simulated_row = results[-1:]
# print(simulated_row)
# new_dataframe = pd.DataFrame(simulated_row)
# new_dataframe.columns = ['time', 'glucose', 'serine', 'biomass', 'mu']
# data_frame = data_frame.append(new_dataframe, ignore_index=True)
# print(selected_data['CO2 (Vol.%)'][0:-1],'WOW')
# selected_data['CO2 (Vol.%)'].reset_index(inplace=True, drop=True)
# print(selected_data['CO2 (Vol.%)'])
# data_frame['CO2 (%)'] = selected_data['CO2 (Vol.%)'].iloc[0:-1]
#
# print(data_frame)

    # wb = load_workbook("output.xlsx")
    # ws = wb['Sheet1']
    # row = list(data_frame.iloc[-1])
    # ws.append(row)
    # wb.save("output.xlsx")
    #
    # output_values = 'output.xlsx'
    #
    # output_values = pd.ExcelFile(output_values)
    #
    # # Loads the sheet we want to work with
    # data_frame = output_values.parse('Sheet1')
    #
    # print(range(0, (len(mu) - 2)))
#data_frame = []


# for i in range(0, (len(mu) - 1)):
#     print(data_frame, "on the top")
#     #print(results,"on the top")
#     #print(data_frame['glucose'].iloc[-1])
#     #r.reset()
#     print(mu[i+1],'mu')
#     r.mu = mu[i+1]
#     glucose = data_frame['glucose'].iloc[-1]
#     serine = data_frame['serine'].iloc[-1]
#     biomass = data_frame['biomass'].iloc[-1]
#     print(glucose,serine,biomass)
#
#     alpha_online, beta_online = parameter_estimation_online(filename_experimental_data1,
#                                                             filename_experimental_data2,
#                                                             alpha_lower_bound, alpha_upper_bound,
#                                                             beta_lower_bound, beta_upper_bound,
#                                                             str(mu[i+1]), str(glucose), str(serine),
#                                                             str(biomass))
#     #print((mu[i+1]))
#     print(alpha_online,beta_online)
#     r.alpha = float(alpha_online)
#     r.beta = float(beta_online)
#     start_time = selected_time_decimals_hours[i]
#     end_time = selected_time_decimals_hours[i + 1]
#     results = r.simulate(start_time, end_time, 2)
#     print(results)
#     simulated_row = results[-1:]
#     print(simulated_row)
#     new_dataframe = pd.DataFrame(simulated_row)
#     new_dataframe.columns = ['time', 'glucose', 'serine', 'biomass', 'mu']
#     data_frame = data_frame.append(new_dataframe, ignore_index=True)
#     print(data_frame)
#     #print("ROUND",i)
#
# print(data_frame)
    # print(selected_data['CO2 (Vol.%)'])


#
# trace1 = go.Scatter(
#     x=selected_time_decimals_hours,
#     y=selected_data['CO2 (Vol.%)'],
#     name='CO2'
# )
#
# trace2 = go.Scatter(
#     x=selected_time_decimals_hours,
#     y=mu,
#     name='mu'
# )
#
# trace3 = go.Scatter(
#     x=data_frame['time'],
#     y=data_frame['mu'],
#     name='mu'
# )
#
# trace4 = go.Scatter(
#     x=data_frame['time'],
#     y=data_frame['biomass'],
#     name='Biomass'
# )
#
# trace5 = go.Scatter(
#     x=data_frame['time'],
#     y=data_frame['serine'],
#     name='Serine'
# )
#
# trace6 = go.Scatter(
#     x=data_frame['time'],
#     y=data_frame['glucose'],
#     name='Glucose'
# )
#
# fig = tools.make_subplots(rows=2, cols=3, subplot_titles=('CO2 online data', 'mu from CO2',
#                                                           'mu from model', 'Biomass from model',
#                                                           'Serine from model', 'Glucose from model'))
#
# fig.append_trace(trace1, 1, 1)
# fig.append_trace(trace2, 1, 2)
# fig.append_trace(trace3, 1, 3)
# fig.append_trace(trace4, 2, 1)
# fig.append_trace(trace5, 2, 2)
# fig.append_trace(trace6, 2, 3)
#
# fig['layout'].update(height=640, width=1260,
#                      margin=dict(
#                          l=120,
#                          r=100,
#                          b=100,
#                          t=70,
#                          pad=2
#                      ))
#
# fig['layout']['yaxis1'].update(showgrid=True, title='CO2 (%)', exponentformat='power', nticks=10,
#                                tickfont=dict(size=10), domain=[0.65, 1])
# fig['layout']['yaxis2'].update(showgrid=True, title='Mu (1/h)', exponentformat='power', nticks=10,
#                                tickfont=dict(size=10), domain=[0.65, 1])
# fig['layout']['yaxis3'].update(showgrid=True, title='Mu (1/h)', exponentformat='power', nticks=10,
#                                tickfont=dict(size=10), domain=[0.65, 1])
# fig['layout']['yaxis4'].update(showgrid=True, title='Biomass (moles)', exponentformat='power', nticks=10,
#                                tickfont=dict(size=10), domain=[0, 0.35])
# fig['layout']['yaxis5'].update(showgrid=True, title='Serine (moles)', exponentformat='power', nticks=10,
#                                tickfont=dict(size=10), domain=[0, 0.35])
# fig['layout']['yaxis6'].update(showgrid=True, title='Glucose (moles)', exponentformat='power', nticks=10,
#                                tickfont=dict(size=10), domain=[0, 0.35])
#
# fig['layout']['xaxis1'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
#                                domain=[0, 0.27])
# fig['layout']['xaxis2'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
#                                domain=[0.36, 0.63])
# fig['layout']['xaxis3'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
#                                domain=[0.72, 0.99])
# fig['layout']['xaxis4'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
#                                domain=[0, 0.27])
# fig['layout']['xaxis5'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
#                                domain=[0.36, 0.63])
# fig['layout']['xaxis6'].update(showgrid=True, title='Time (hours)', nticks=10, tickfont=dict(size=10),
#                                domain=[0.72, 0.99])
#
# plotly.offline.plot(fig)

######################################

# import dash
# from dash.dependencies import Input, Output, State
# import dash_core_components as dcc
# import dash_html_components as html
# import dash_table_experiments as dt
# import json
# import pandas as pd
# import numpy as np
# import plotly
#
# app = dash.Dash()
#
# app.scripts.config.serve_locally = True
# # app.css.config.serve_locally = True
#
# DF_WALMART = pd.read_csv('https://raw.githubusercontent.com/plotly/datasets/master/1962_2006_walmart_store_openings.csv')
#
# DF_GAPMINDER = pd.read_csv(
#     'https://raw.githubusercontent.com/plotly/datasets/master/gapminderDataFiveYear.csv'
# )
# DF_GAPMINDER = DF_GAPMINDER[DF_GAPMINDER['year'] == 2007]
# DF_GAPMINDER.loc[0:20]
#
# DF_SIMPLE = pd.DataFrame({
#     'x': ['A', 'B', 'C', 'D', 'E', 'F'],
#     'y': [4, 3, 1, 2, 3, 6],
#     'z': ['a', 'b', 'c', 'a', 'b', 'c']
# })
#
# ROWS = [
#     {'a': 'AA', 'b': 1},
#     {'a': 'AB', 'b': 2},
#     {'a': 'BB', 'b': 3},
#     {'a': 'BC', 'b': 4},
#     {'a': 'CC', 'b': 5},
#     {'a': 'CD', 'b': 6}
# ]
#
#
# app.layout = html.Div([
#     html.H4('Gapminder DataTable'),
#     dt.DataTable(
#         rows=DF_GAPMINDER.to_dict('records'),
#
#         # optional - sets the order of columns
#         columns=sorted(DF_GAPMINDER.columns),
#
#         row_selectable=True,
#         filterable=True,
#         sortable=True,
#         selected_row_indices=[],
#         id='datatable-gapminder'
#     ),
#     html.Div(id='selected-indexes'),
#     dcc.Graph(
#         id='graph-gapminder'
#     ),
# ], className="container")
#
#
# @app.callback(
#     Output('datatable-gapminder', 'selected_row_indices'),
#     [Input('graph-gapminder', 'clickData')],
#     [State('datatable-gapminder', 'selected_row_indices')])
# def update_selected_row_indices(clickData, selected_row_indices):
#     if clickData:
#         for point in clickData['points']:
#             if point['pointNumber'] in selected_row_indices:
#                 selected_row_indices.remove(point['pointNumber'])
#             else:
#                 selected_row_indices.append(point['pointNumber'])
#     return selected_row_indices
#
#
# @app.callback(
#     Output('graph-gapminder', 'figure'),
#     [Input('datatable-gapminder', 'rows'),
#      Input('datatable-gapminder', 'selected_row_indices')])
# def update_figure(rows, selected_row_indices):
#     dff = pd.DataFrame(rows)
#     fig = plotly.tools.make_subplots(
#         rows=3, cols=1,
#         subplot_titles=('Life Expectancy', 'GDP Per Capita', 'Population',),
#         shared_xaxes=True)
#     marker = {'color': ['#0074D9']*len(dff)}
#     for i in (selected_row_indices or []):
#         marker['color'][i] = '#FF851B'
#     fig.append_trace({
#         'x': dff['country'],
#         'y': dff['lifeExp'],
#         'type': 'bar',
#         'marker': marker
#     }, 1, 1)
#     fig.append_trace({
#         'x': dff['country'],
#         'y': dff['gdpPercap'],
#         'type': 'bar',
#         'marker': marker
#     }, 2, 1)
#     fig.append_trace({
#         'x': dff['country'],
#         'y': dff['pop'],
#         'type': 'bar',
#         'marker': marker
#     }, 3, 1)
#     fig['layout']['showlegend'] = False
#     fig['layout']['height'] = 800
#     fig['layout']['margin'] = {
#         'l': 40,
#         'r': 10,
#         't': 60,
#         'b': 200
#     }
#     fig['layout']['yaxis3']['type'] = 'log'
#     return fig
#
#
# app.css.append_css({
#     'external_url': 'https://codepen.io/chriddyp/pen/bWLwgP.css'
# })
#
# if __name__ == '__main__':
#     app.run_server(debug=True)